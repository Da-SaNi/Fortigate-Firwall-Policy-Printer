# encoding=utf8

# pyinstaller -w -F -i="./avatar.ico" "Fortigate Firewall Policy Printer(Notebook).py"

import requests
import logging
import openpyxl
import pandas as pd
import datetime
import sys
reload(sys)
sys.setdefaultencoding('utf8')
from openpyxl.styles import Border, Side, Alignment, PatternFill
import urllib3.exceptions
from openpyxl.utils.dataframe import dataframe_to_rows
import urllib
import wx

class FortiGate:
    def __init__(self, ipaddr, username, password, vdom, timeout=10, port="443", verify=False):
        self.ipaddr = ipaddr
        self.username = username
        self.password = urllib.quote(password)
        self.port = port
        self.urlbase = "https://{ipaddr}:{port}/".format(ipaddr=self.ipaddr,port=self.port)
        self.timeout = timeout
        self.vdom = vdom
        self.verify = verify


    # Login / Logout Handlers
    def login(self):
        """
        Log in to FortiGate with info provided in during class instantiation
        :return: Open Session
        """
        session = requests.session()
        if not self.verify:
            # Disable requests' warnings for insecure connections
            urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

        url = self.urlbase + 'logincheck'

        # Login
        session.post(url,
                     data='username={username}&secretkey={password}'.format(username=self.username,
                                                                            password=self.password),
                     verify=self.verify,
                     timeout=self.timeout)

        # Get CSRF token from cookies, add to headers
        for cookie in session.cookies:
            if cookie.name == 'ccsrftoken':
                csrftoken = cookie.value[1:-1]  # strip quotes
                session.headers.update({'X-CSRFTOKEN': csrftoken})

        # Check whether login was successful
        login_check = session.get(self.urlbase + "api/v2/cmdb/system/vdom")
        login_check.raise_for_status()
        return session

    def logout(self, session):
        """
        Log out of device
        :param session: Session created by login method
        :return: None
        """
        url = self.urlbase + 'logout'
        session.get(url, verify=self.verify, timeout=self.timeout)
        logging.info("Session logged out.")

    # General Logic Methods
    def does_exist(self, object_url):
        """
        GET URL to assert whether it exists within the firewall
        :param object_url: Object to locate
        :return: Bool - True if exists, False if not
        """
        session = self.login()
        request = session.get(object_url, verify=self.verify, timeout=self.timeout, params='vdom='+self.vdom)
        self.logout(session)
        if request.status_code == 200:
            return True
        else:
            return False

    # API Interaction Methods
    def get(self, url):
        """
        Perform GET operation on provided URL
        :param url: Target of GET operation
        :return: Request result if successful (type list), HTTP status code otherwise (type int)
        """
        session = self.login()
        request = session.get(url, verify=self.verify, timeout=self.timeout, params='vdom='+self.vdom)
        self.logout(session)
        if request.status_code == 200:
            return request.json()
        else:
            return request.status_code

    def get_firewall_policy(self, specific=False, filters=False):
        """
        Get firewall policy information from firewall
        :param specific: If provided, a specific object will be returned. If not, all objects will be returned.
            Specific can either be the policy name, or the policy ID.
        :param filters: If provided, the raw filter is appended to the API call.
        :return: JSON data for all objects in scope of request, nested in a list.
        """
        api_url = self.urlbase + "api/v2/cmdb/firewall/policy?vdom=" + self.vdom
        if specific:
            if type(specific) == int:
                api_url += str(specific)
            else:
                api_url += "?filter=name==" + specific
        elif filters:
            api_url += "?filter=" + filters
        results = self.get(api_url)  # get requestpolicy_id
        if type(results) == int:
            return results
        elif len(results) == 0:
            return 404
        else:
            return results

    def parsing(self):
        json_data = self.get_firewall_policy()
        results = json_data.get("results")

        srcintf = []
        dstintf = []
        srcaddr = []
        dstaddr = []
        service = []
        poolname = []

        data_results = []
        dataframe = []

        for x in range(0, len(results)):

            for X in range(0, len(results[x]["srcintf"])):
                srcintf.append(results[x]["srcintf"][X]['q_origin_key'])
            for X in range(0, len(results[x]["dstintf"])):
                dstintf.append(results[x]["dstintf"][X]['q_origin_key'])
            for X in range(0, len(results[x]["srcaddr"])):
                srcaddr.append(results[x]["srcaddr"][X]['q_origin_key'])
            for X in range(0, len(results[x]["dstaddr"])):
                dstaddr.append(results[x]["dstaddr"][X]['q_origin_key'])
            for X in range(0, len(results[x]["service"])):
                service.append(results[x]["service"][X]['q_origin_key'])
            try:
                for X in range(0, len(results[x]["poolname"])):
                    poolname.append(results[x]["poolname"][X]['q_origin_key'])
            except:
                poolname.append(u"")

            data_results.append(results[x]["policyid"])
            data_results.append(results[x]["name"])
            data_results.append(' '.join(srcintf))
            data_results.append(' '.join(dstintf))
            data_results.append(' '.join(srcaddr))
            data_results.append(' '.join(dstaddr))
            data_results.append(' '.join(service))
            data_results.append(results[x]["action"])
            data_results.append(results[x]["status"])
            data_results.append(results[x]["schedule"])
            data_results.append(results[x]["logtraffic"])
            data_results.append(results[x]["nat"])
            data_results.append(results[x]["ippool"])
            data_results.append(' '.join(poolname))
            data_results.append(results[x]["comments"])
            srcintf = []
            dstintf = []
            srcaddr = []
            dstaddr = []
            service = []
            poolname = []
            dataframe.append(data_results)
            data_results = []
        return dataframe

    def generate_xlsx(self, results, fdName):
        raw_data = []
        d = datetime.datetime.now()
        keys = ['Policy_ID', 'Name', 'Src_Intf', 'Dst_Intf', 'Src_Addr', 'Dst_Addr', 'Service', 'Action', 'Status',
                'Schedule', 'Log-Traffic', 'NAT', 'IP_POOL', 'Pool_Name', 'Comments']
        fName = fdName

        implicitDeny = ["0", "implicit_Deny", "all", "all", "all", "all", "Deny", "", "always", "all", "Disabled", "",
                        "",
                        "", ""]

        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        raw_data = results
        raw_data.insert(0, keys)
        raw_data.append(implicitDeny)

        df = pd.DataFrame.from_records(raw_data)

        wb = openpyxl.Workbook()
        ws = wb.worksheets[0]
        ws.title = u"방화벽 정책"

        # ws 시트에 df를 입력
        for r in dataframe_to_rows(df, index=False, header=False):
            ws.append(r)

        # 음영지정
        for j in range(ws.max_column):
            ws.cell(row=1, column=j + 1).fill = PatternFill(start_color='808080', end_color='808080', fill_type='solid')

        for j in range(ws.max_column):
            ws.cell(row=ws.max_row, column=j + 1).fill = PatternFill(start_color='808080', end_color='808080',
                                                                     fill_type='solid')

        # 개행문자 사용을 위한 wrap_text 지정
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical='top')

        # 공백을 줄바꿈으로 전환
        for col_num in range(1, ws.max_column + 1):
            for row_num in range(1, ws.max_row + 1):

                # 활성화 된 셀에 테두리
                ws.cell(row=row_num, column=col_num).border = thin_border
                # tempstr : cell값이 문자열이 아닌 경우를 감안하여 str로 바꿔줌
                # print(ws.cell(row=row_num, column=col_num).value)
                if isinstance(ws.cell(row=row_num, column=col_num).value, int) == True:
                    tempstr = str(ws.cell(row=row_num, column=col_num).value)
                else:
                    tempstr = str(ws.cell(row=row_num, column=col_num).value)

                # 문자열 함수 replace 사용(,를 개행문자로)
                data = tempstr.replace(" ", "\n")

                # 빈 셀의 경우 None이라는 문자열 타입이므로 제외하고 데이터 입력
                if data != "None":
                    ws.cell(row=row_num, column=col_num).value = data

                    data = ""

        # 열 너비 자동 맞춤
        columns = None

        for i, column_cells in enumerate(ws.columns):
            is_ok = False
            if columns == None:
                is_ok = True
            elif isinstance(columns, list) and i in columns:
                is_ok = True

            if is_ok:
                length = max(str(cell.value).find("\n") for cell in column_cells)
                if length == -1:
                    length = max(len(str(cell.value)) for cell in column_cells)
                ws.column_dimensions[column_cells[0].column_letter].width = length + 3

        # 엑셀 파일 저장
        wb.save(fName)
        return None

class window(wx.Frame):
    def __init__(self):
        wx.Frame.__init__(self, parent=None, title="Fortigate Policy Printer", size=wx.Size(170, 450))
        self.mainPanel = wx.Panel(self)
        
        self.else_label = wx.StaticText(self.mainPanel, -1, "Please Input IP", pos=(20, 110))

        self.zone_label = wx.StaticText(self.mainPanel, -1, "[ Firewall ]", pos=(20, 130))
        self.zone_text = wx.TextCtrl(self.mainPanel, -1, "1", pos=(20, 150))
        self.zone_text.Bind(wx.EVT_TEXT_ENTER, self.retnvalue)

        self.id_label = wx.StaticText(self.mainPanel, -1, "[ ID ]", pos=(20, 180))
        self.id_text = wx.TextCtrl(self.mainPanel, -1, "", pos=(20, 200))
        self.id_text.Bind(wx.EVT_TEXT_ENTER, self.retnvalue)

        self.pw_label = wx.StaticText(self.mainPanel, -1, "[ PW ]", pos=(20, 230))
        self.pw_text = wx.TextCtrl(self.mainPanel, -1, "", style=wx.TE_PASSWORD, pos=(20, 250))
        self.pw_text.Bind(wx.EVT_TEXT_ENTER, self.retnvalue)

        self.vdom_label = wx.StaticText(self.mainPanel, -1, "[ VDOM ]", pos=(20, 280))
        self.vdom_text = wx.TextCtrl(self.mainPanel, -1, "", pos=(20, 300))
        self.vdom_text.Bind(wx.EVT_TEXT_ENTER, self.retnvalue)

        self.btn_click = wx.Button(self.mainPanel, -1, label="RUN!", pos=(20, 340), size=(115, 30))
        self.btn_click.Bind(wx.EVT_BUTTON, self.retnvalue)

    def retnvalue(self, event):
        self.firewall = self.zone_text.GetValue()
        self.id = self.id_text.GetValue()
        self.pw = self.pw_text.GetValue()
        self.nc_num = self.vdom_text.GetValue()
        d = datetime.datetime.now()
        file_name = u'FirewallPolicy_' + d.strftime("%Y") + d.strftime("%m") + d.strftime("%d") + u'.xlsx'

        if self.id == "":
            wx.MessageBox(u"ID를 입력하세요.", 'Warning', wx.OK)
        elif self.pw == "":
            wx.MessageBox(u"패스워드를 입력하세요.", 'Warning', wx.OK)
        elif self.nc_num == "":
            wx.MessageBox(u"NC정보를 입력하세요.", 'Warning', wx.OK)

        if self.firewall != "" and self.nc_num != "" and self.id != "" and self.pw != "":
            try:
                connect = FortiGate(ipaddr=self.firewall, username=self.id, password=self.pw, vdom=self.nc_num)
                session = connect.login()
                connect.generate_xlsx(results=connect.parsing(), fdName=file_name)
                connect.logout(session)
                wx.MessageBox(u'출력 완료!', 'Info', wx.OK)
            except Exception:
                wx.MessageBox(u'출력 실패!', 'Info', wx.OK)
                connect.logout(session)

        return 0

if __name__ == "__main__":
    app = wx.App()
    frame = window()
    frame.Show()
    app.MainLoop()
